import time
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import sys
from openpyxl.styles import PatternFill, Color, Border, Side
from mapping import (
    title_map, configuration_map, folder_map, tape_map, not_monitored_map,
    kasp_excel_unhosted_map, kasp_excel_hosted_map, kasp_hosted_move_map
)
import win32com.client
import os
import traceback
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


def move_email(mail, folder_map, backup_folder, mail_sender):
    """Moves email to the appropriate folder based on the sender."""
    mail.UnRead = False
    move_map = folder_map.get(mail_sender)
    if move_map is not None:
        move_folder = backup_folder.Folders[move_map]
        mail.Move(move_folder)
        time.sleep(1)
    else:
        print(f"No folder found for sender: {mail_sender}")


def import_data(column, mail_result):
    """Imports backup check results into the corresponding Excel cell."""
    col = get_column_letter(column)
    col1 = get_column_letter(column + 1)
    if mail_result == "[Success]":
        sheet_copy[f"{col}3"] = "OK"
        sheet_copy[f"{col1}3"] = "AUTO"
    elif mail_result == "[Warning]":
        sheet_copy[f"{col}3"] = "WARNING"
        sheet_copy[f"{col1}3"] = ""
    elif mail_result == "[Failed]":
        sheet_copy[f"{col}3"] = "NOK"
        sheet_copy[f"{col1}3"] = ""


# Initial setup for paths and checking environment
exe_current_path = os.path.dirname(sys.argv[0])
if "DFSRoot" in exe_current_path: #The .exe file was uploaded to the DFSRoot folder. To run it properly, users need to copy it to their PC first, so I included this step to ensure compliance.
    print("You are trying to run the program from DFSRoot. \nPlease copy the executable file to your PC and try again.")
    time.sleep(40)
    sys.exit()

while True:
    print("""Welcome to Backup Auto Check!
Please make sure that Excel files and Citrix Outlook are closed before starting.
Ensure Outlook remains open while the program is running.""")
    start_prompt = input("Type 'backup' to start: ")
    if start_prompt.lower() == "backup":
        break

# Paths to original and copied Excel files
file_path_original = "path to your original file"
original_path = Path(file_path_original)
workbook_original = openpyxl.load_workbook(original_path)
sheet_original = workbook_original["Backup Checking"]

# Check for the correct version in the original file
if sheet_original["A3"].value != "V1.5":
    print("You are using an older version. Please get the latest version.")
    time.sleep(30)
    sys.exit()

# Create or open the copy of the Excel file
desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
file_path_copy = f"{desktop_path}/Backup_data.xlsx"
copy_path = Path(file_path_copy)

if copy_path.exists():
    print(f"Excel file already exists.")
    workbook_copy = openpyxl.load_workbook(copy_path)
    path_exists = True
else:
    workbook_copy = openpyxl.Workbook()
    workbook_copy.save(copy_path)
    print("Excel file created.")
    path_exists = False

sheet_copy = workbook_copy["Sheet"]

# Set constants for column handling
active_cols = 416
merged_cols = 208

# Initialize copy Excel file if it doesn't exist
if not path_exists:
    for i in range(1, active_cols):
        col = get_column_letter(i)
        sheet_copy.column_dimensions[col].width = 20
    sheet_copy.row_dimensions[2].height = 30

    for i in range(1, merged_cols):
        col1 = get_column_letter((i - 1) * 2 + 2)
        col2 = get_column_letter(i * 2 + 1)
        merge_range = f"{col1}2:{col2}2"
        sheet_copy.merge_cells(merge_range)

    # Formatting the Excel file
    new_color = Color(theme=5, tint=0.8, type='theme')
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    new_fill = PatternFill(start_color=new_color, end_color=new_color, fill_type='solid')
    for cell in sheet_copy[2]:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
    for cell in sheet_copy[3]:
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        cell.fill = new_fill
        cell.border = border
    sheet_copy["A3"].fill = PatternFill(fill_type="none")
    sheet_copy["A3"].alignment = openpyxl.styles.Alignment(horizontal='right')

# Check date and title consistency between original and copy files
date_format = "D/M/YYYY"
current_date = datetime.now().date()

if sheet_copy["A3"].value is None:
    sheet_copy["A3"] = current_date
    sheet_date = current_date.year
else:
    sheet_date = sheet_copy["A3"].value.date()

same_date = sheet_date == current_date

same_title = True
for i in range(1, merged_cols):
    col = get_column_letter((i - 1) * 2 + 2)
    if sheet_copy[f"{col}2"].value != sheet_original[f"{col}5"].value:
        same_title = False
        break

if not same_date or not same_title:
    for cell in sheet_copy[3]:
        cell.value = None
    sheet_copy["A3"] = current_date

    if not same_title:
        for i in range(1, merged_cols):
            col = get_column_letter((i - 1) * 2 + 2)
            sheet_copy[f"{col}1"].value = sheet_original[f"{col}4"].value
            sheet_copy[f"{col}2"].value = sheet_original[f"{col}5"].value

sheet_copy["A3"].number_format = date_format
workbook_copy.save(copy_path)
workbook_original.close()

print("Excel is ready to accept data.")

# Initialize email processing variables
save_counter = 0
attempt = 1
kasp_check_alarm = ("Critical", "Warning")
kasp_check_title = ("Kaspersky", "KSC", "KES")
kasp_hosted = ("Customer 1", "Customer 2", "Customer 3")  # Placeholder customer names

# Try connecting to Outlook
while attempt <= 5:
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.Folders.Item("your-email@example.com").Folders.Item("Inbox")
        bits_folder = inbox.Folders.Item("Your inbox sub-folder name")
        Erledigt_folder = outlook.Folders.Item("your-email@example.com").Folders.Item("Folder name")
        backup_folder = Erledigt_folder.Folders.Item("Sub-folder name")

        if bits_folder is None:
            print("BITS folder not found.")

        mailbox = bits_folder.Items
        mailbox.Sort("[ReceivedTime]", True)
        total_mails = mailbox.Count - 1
        i = total_mails
        break
    except:
        print("Couldn't connect to Outlook. Retrying...")
        attempt += 1
        time.sleep(10)

# Email processing loop
while i >= 0:
    try:
        time.sleep(0.1)
        mail = mailbox[i]
        print(f"Processing {mail} ...")
        i -= 1
        excel_title = None
        mail_subject = mail.Subject
        mail_sender = mail.SenderEmailAddress
        mail_list = mail_subject.split()
        attempt = 1

# There were some special categories that needed to be handled differently, some examples can be found below. Add special handling according to your needs, everything dependes on the structure of the mails you need to process.

        # Skip processing if mail list is too short
        if len(mail_list) <= 2:
            continue

        # List of email addresses and IDs not to monitor
        not_monitored = (
            "costumer1@example.com",
            "costumer2@example.com",
            "costumer3@example.com",
            "costumer4@example.com"
        )

        if mail_sender in not_monitored:
            move_email(mail, not_monitored_map, backup_folder, mail_sender)
            continue

        # -------------------Clear title------------------------------------------------------------
        clear_up_to = ("objects)", "objects),", "machines)", "VMs)")
        clear_up_to_index = None
        for trigger in clear_up_to:
            try:
                clear_up_to_index = mail_list.index(trigger) - 1
            except ValueError:
                continue

        mail_result = mail_list[0]
        if clear_up_to_index is None:
            mail_title = " ".join(mail_list[1:])
        else:
            mail_title = " ".join(mail_list[1:clear_up_to_index])
        # -------------------Clear title end-----------------------------------------------------------

        if mail_list[1] == "Manually" and mail_result == "[Success]": #Special category that we didn't had to import the data on excel
            move_email(mail, folder_map, backup_folder, mail_sender)
            continue

        if mail_list[2] == "Configuration": #Special category where mail subject was identical for all costumers so it uses sender's email address to handle them.
            excel_title_config = configuration_map.get(mail_sender)
            if excel_title_config is not None:
                for column in range(1, active_cols):
                    cell_title = sheet_copy.cell(row=2, column=column).value
                    if cell_title == excel_title_config:
                        import_data(column, mail_result)
                        if mail_result == "[Success]":
                            move_email(mail, folder_map, backup_folder, mail_sender)
                        break
            continue

        if mail_title == "Backup to Tape Job 1": #Special category where mail subject was identical for all costumers so it uses sender's email address to handle them.
            excel_title_tape = tape_map.get(mail_sender)
            for column in range(1, active_cols):
                cell_title = sheet_copy.cell(row=2, column=column).value
                if cell_title == excel_title_tape:
                    import_data(column, mail_result)
                    if mail_result == "[Success]":
                        move_email(mail, folder_map, backup_folder, mail_sender)
                    break
            continue

        if "NASSPSO01" in mail_title:  # Non-monitored backup
            move_email(mail, folder_map, backup_folder, mail_sender)
            continue

        # -------------------------------------- KASPERSKY START -------------------------------------------------------

        for word in kasp_check_title:
            if word in mail_subject:
                mail_body = mail.Body
                kasp_is_clear = all(alarm_keyword not in mail_body for alarm_keyword in kasp_check_alarm)


                if mail_sender in not_monitored:
                    move_email(mail, folder_map, backup_folder, mail_sender)

                if kasp_is_clear:
                    if mail_sender != "special@example.com":     # Exchange's internal systems
                        excel_title = kasp_excel_unhosted_map.get(mail_sender)
                        if excel_title is not None:
                            for column in range(1, active_cols):
                                cell_title = sheet_copy.cell(row=2, column=column).value
                                if cell_title == excel_title:
                                    col = get_column_letter(column)
                                    col1 = get_column_letter(column + 1)
                                    sheet_copy[f"{col}3"] = "OK"
                                    sheet_copy[f"{col1}3"] = "AUTO"
                            move_email(mail, folder_map, backup_folder, mail_sender)

                    elif mail_sender == "special@example.com":  #Emails from regular customer email addresses
                        costumer = None
                        for possible_costumer in kasp_hosted:
                            if possible_costumer in mail_subject:
                                costumer = possible_costumer
                                print(costumer)
                                break

                        if costumer is not None:
                            excel_title = kasp_excel_hosted_map.get(costumer)
                            for column in range(1, active_cols):
                                cell_title = sheet_copy.cell(row=2, column=column).value
                                if cell_title == excel_title:
                                    col = get_column_letter(column)
                                    col1 = get_column_letter(column + 1)
                                    sheet_copy[f"{col}3"] = "OK"
                                    sheet_copy[f"{col1}3"] = "AUTO"
                                    break
                            move_email(mail, kasp_hosted_move_map, backup_folder, costumer)
                    else:
                        print("Cannot identify customer")
                else:
                    break

        # -------------------------------------- KASPERSKY END -----------------------------------------------------------------

        excel_title = title_map.get(mail_title)

        if excel_title is None: #if it can't find the excel title in any dictionary it will ignore the mail and continue
            print(f"Processing {mail} Done")
            continue

        for column in range(1, active_cols):  #Handler for normal emails
            cell_title = sheet_copy.cell(row=2, column=column).value
            if cell_title == excel_title:
                import_data(column, mail_result)
                if mail_result == "[Success]":
                    move_email(mail, folder_map, backup_folder, mail_sender)
                break
        print(f"Processing {mail} Done")

    except Exception as e:
        with open('error_log.txt', 'a') as f:
            f.write('Exception occurred: \n')
            traceback.print_exc(file=f)
            print("Error occurred. Retrying...")
            while attempt <= 5:
                try:
                    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
                    inbox = outlook.Folders.Item("your_email@example.com").Folders.Item("Inbox")
                    bits_folder = inbox.Folders.Item("BITS")
                    Erledigt_folder = outlook.Folders.Item("your_email@example.com").Folders.Item("_- Erledigt -_")
                    backup_folder = Erledigt_folder.Folders.Item("_BITS Backup & Kaspersky Kontrolle")
                    mailbox = bits_folder.Items
                    mailbox.Sort("[ReceivedTime]", True)
                    break
                except:
                    print(f"Attempt {attempt}: Error occurred. Retrying...")
                    attempt += 1
                    time.sleep(10)
                    continue
            else:
                print("5 retries failed. Program will stop!")
                break

    save_counter += 1
    if save_counter >= 5:
        workbook_copy.save(copy_path)
        save_counter = 0
        print("SAVED!")

workbook_copy.save(copy_path)
workbook_copy.close()
print("Job completed successfully! \n\nYou can copy the imported data to the original Excel.")
time.sleep(30)